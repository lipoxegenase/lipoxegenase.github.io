#!/usr/bin/env python3
"""
KatalystVC Excel API Microservice
A Flask-based API for capturing website form submissions and storing them in an Excel file.
"""

import os
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configuration
EXCEL_FILE = 'katalystvc_leads.xlsx'
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
EMAIL_USER = os.getenv('EMAIL_USER', 'support@katalystvc.com')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD', '')  # App password for Gmail
INTERNAL_EMAIL = os.getenv('INTERNAL_EMAIL', 'support@katalystvc.com')

# Excel column headers
EXCEL_HEADERS = [
    'Timestamp', 'Lead ID', 'First Name', 'Last Name', 'Email', 'Company', 
    'Role', 'Phone', 'Topic', 'Notes', 'Consent', 'Source Page', 
    'UTM Source', 'UTM Medium', 'UTM Campaign', 'UTM Term', 'UTM Content'
]

def initialize_excel_file():
    """Initialize the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Leads"
        
        # Add headers
        for col_num, header in enumerate(EXCEL_HEADERS, 1):
            worksheet.cell(row=1, column=col_num, value=header)
        
        # Auto-adjust column widths
        for col_num, header in enumerate(EXCEL_HEADERS, 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header) + 2, 15)
        
        workbook.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

def get_next_lead_id():
    """Get the next available lead ID by checking the last row in the Excel file."""
    try:
        workbook = load_workbook(EXCEL_FILE)
        worksheet = workbook.active
        
        # Find the last row with data
        last_row = worksheet.max_row
        if last_row <= 1:  # Only headers exist
            return 1
        
        # Get the last lead ID and increment
        last_lead_id = worksheet.cell(row=last_row, column=2).value  # Lead ID is in column 2
        return (last_lead_id or 0) + 1
    except Exception as e:
        print(f"Error getting next lead ID: {e}")
        return 1

def append_to_excel(lead_data):
    """Append lead data to the Excel file."""
    try:
        workbook = load_workbook(EXCEL_FILE)
        worksheet = workbook.active
        
        # Find the next empty row
        next_row = worksheet.max_row + 1
        
        # Prepare data row
        data_row = [
            lead_data.get('timestamp'),
            lead_data.get('lead_id'),
            lead_data.get('first_name'),
            lead_data.get('last_name'),
            lead_data.get('email'),
            lead_data.get('company'),
            lead_data.get('role'),
            lead_data.get('phone'),
            lead_data.get('topic'),
            lead_data.get('notes'),
            lead_data.get('consent'),
            lead_data.get('source_page'),
            lead_data.get('utm_source'),
            lead_data.get('utm_medium'),
            lead_data.get('utm_campaign'),
            lead_data.get('utm_term'),
            lead_data.get('utm_content')
        ]
        
        # Write data to the worksheet
        for col_num, value in enumerate(data_row, 1):
            worksheet.cell(row=next_row, column=col_num, value=value)
        
        # Save the workbook
        workbook.save(EXCEL_FILE)
        print(f"Successfully added lead {lead_data.get('lead_id')} to Excel file")
        return True
        
    except Exception as e:
        print(f"Error writing to Excel file: {e}")
        return False

def send_email(to_email, subject, body, is_html=False):
    """Send an email using SMTP."""
    if not EMAIL_PASSWORD:
        print("Email credentials not configured. Skipping email send.")
        return False
    
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = EMAIL_USER
        msg['To'] = to_email
        msg['Subject'] = subject
        
        if is_html:
            msg.attach(MIMEText(body, 'html'))
        else:
            msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_USER, to_email, text)
        server.quit()
        return True
    except Exception as e:
        print(f"Email sending failed: {e}")
        return False

def get_confirmation_email_template(first_name, topic):
    """Generate confirmation email for the user."""
    topic_mapping = {
        'infra': 'AI-Ready Infrastructure Audit',
        'fhir': 'FHIR/TEFCA 90-Day Sprint'
    }
    
    topic_display = topic_mapping.get(topic, topic)
    download_link = f"https://katalystvc.com/downloads/{topic}-brief.html"
    
    subject = "Your KatalystVC Inquiry Has Been Received - What's Next?"
    
    body = f"""Dear {first_name},

Thank you for reaching out to KatalystVC. We've successfully received your inquiry regarding {topic_display}.

We appreciate your interest in our specialized technical consulting services. Our team is currently reviewing your submission and will get back to you within 24-48 business hours to schedule your 20-minute diagnostic session.

In the meantime, you can access your requested marketing brief here:
{download_link}

We look forward to connecting with you and exploring how KatalystVC can help you achieve your technical objectives with tight, efficient, and smart solutions.

Best regards,

The KatalystVC Team
support@katalystvc.com"""
    
    return subject, body

def get_internal_notification_template(lead_data):
    """Generate internal notification email for the team."""
    subject = f"NEW KatalystVC Lead: {lead_data['first_name']} {lead_data['last_name']} - {lead_data['topic']}"
    
    body = f"""Hello Team,

A new lead has been submitted via the KatalystVC website. Please find the details below:

**Lead Information:**
• Lead ID: {lead_data['lead_id']}
• Name: {lead_data['first_name']} {lead_data['last_name']}
• Email: {lead_data['email']}
• Company: {lead_data.get('company', 'N/A')}
• Role: {lead_data.get('role', 'N/A')}
• Phone: {lead_data.get('phone', 'N/A')}
• Topic of Interest: {lead_data['topic']}
• Notes/Message: {lead_data.get('notes', 'N/A')}
• Consent to Contact: {'Yes' if lead_data['consent'] else 'No'}

**Tracking Information:**
• Submission Date: {lead_data['timestamp']}
• Source Page: {lead_data.get('source_page', 'N/A')}
• UTM Source: {lead_data.get('utm_source', 'N/A')}
• UTM Medium: {lead_data.get('utm_medium', 'N/A')}
• UTM Campaign: {lead_data.get('utm_campaign', 'N/A')}
• UTM Term: {lead_data.get('utm_term', 'N/A')}
• UTM Content: {lead_data.get('utm_content', 'N/A')}

**Action Required:**
• Follow up with the lead within 24-48 business hours.
• Lead data has been saved to {EXCEL_FILE}

Best regards,

KatalystVC Automated System"""
    
    return subject, body

@app.route('/submit-lead', methods=['POST'])
def submit_lead():
    """Handle lead submission from the website form."""
    data = request.get_json()

    # Validate required fields
    required_fields = ['firstName', 'lastName', 'email', 'topic', 'consent']
    if not all(field in data for field in required_fields):
        return jsonify({'error': 'Missing required fields'}), 400

    try:
        # Generate lead ID and timestamp
        lead_id = get_next_lead_id()
        timestamp = datetime.now().isoformat()
        
        # Prepare lead data for Excel
        lead_data = {
            'timestamp': timestamp,
            'lead_id': lead_id,
            'first_name': data['firstName'],
            'last_name': data['lastName'],
            'email': data['email'],
            'company': data.get('company'),
            'role': data.get('role'),
            'phone': data.get('phone'),
            'topic': data['topic'],
            'notes': data.get('notes'),
            'consent': 'Yes' if data['consent'] else 'No',
            'source_page': data.get('sourcePage'),
            'utm_source': data.get('utmSource'),
            'utm_medium': data.get('utmMedium'),
            'utm_campaign': data.get('utmCampaign'),
            'utm_term': data.get('utmTerm'),
            'utm_content': data.get('utmContent')
        }

        # Write to Excel file
        if not append_to_excel(lead_data):
            return jsonify({'error': 'Failed to save lead data'}), 500

        # Send emails if credentials are configured
        if EMAIL_PASSWORD:
            # Send confirmation email to user
            conf_subject, conf_body = get_confirmation_email_template(data['firstName'], data['topic'])
            user_email_sent = send_email(data['email'], conf_subject, conf_body)
            
            # Send internal notification email
            int_subject, int_body = get_internal_notification_template(lead_data)
            internal_email_sent = send_email(INTERNAL_EMAIL, int_subject, int_body)
            
            print(f"Lead {lead_id} submitted. User email sent: {user_email_sent}, Internal email sent: {internal_email_sent}")
        else:
            print(f"Lead {lead_id} submitted. Email credentials not configured - emails not sent.")

        return jsonify({
            'message': 'Lead submitted successfully',
            'lead_id': lead_id,
            'timestamp': timestamp
        }), 200
        
    except Exception as e:
        print(f"Error processing lead submission: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'excel_file_exists': os.path.exists(EXCEL_FILE)
    }), 200

@app.route('/stats', methods=['GET'])
def get_stats():
    """Get basic statistics about the leads."""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'total_leads': 0, 'excel_file_exists': False}), 200
        
        workbook = load_workbook(EXCEL_FILE)
        worksheet = workbook.active
        total_leads = worksheet.max_row - 1  # Subtract header row
        
        return jsonify({
            'total_leads': max(0, total_leads),
            'excel_file_exists': True,
            'excel_file': EXCEL_FILE
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("KatalystVC Excel API Microservice starting...")
    print(f"Excel file: {EXCEL_FILE}")
    print(f"Email configured: {'Yes' if EMAIL_PASSWORD else 'No (set EMAIL_PASSWORD environment variable)'}")
    
    # Initialize Excel file
    initialize_excel_file()
    
    # Start the Flask application
    app.run(debug=True, host='0.0.0.0', port=5000)
